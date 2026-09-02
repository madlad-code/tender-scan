#!/usr/bin/env python3
"""Beskriv de utlämnade filerna, så att en loader kan skrivas mot dem.

## Varför strukturen och inte innehållet

De nio filerna väger drygt 100 MB. Frestelsen är att låta en modell läsa dem
och sammanfatta — det är fel av två skäl. Det får inte plats, och även om det
fick det vore en modell som ögnar 400 000 fakturarader mindre pålitlig än en
`SUM()` över samma rader.

Modellen ska skriva koden. Koden ska läsa datat. För att skriva koden behövs
inte datat, bara dess form: vilka kolumner som finns, var rubrikraden ligger,
vilken kolumn som är datum och vilket spann den täcker, hur belopp är
formaterade, och om leverantörens organisationsnummer finns med. Det är några
kilobyte per fil i stället för trettio megabyte.

Den här filen tar fram exakt den beskrivningen. Utdatat är litet nog att
klistra in i en session, och det som står i det — leverantörsnamn, belopp — är
utlämnade allmänna handlingar, alltså redan offentligt.

## Vad som är den avgörande frågan per fil

Om varje rad har ett eget fakturadatum. `to_payments` i `payments/base.py`
kastar varje rad utan datum, så en aggregerad "översikt" utan datumkolumn är
obrukbar för utnyttjandegrad hur komplett den än ser ut. Bjurholms
`Leverantörsreskontraöversikt.xlsx` är misstänkt på just den punkten, och
profilen svarar på det.

## Körning

    pip install -e ".[files]"
    python3 scripts/profilera_svarsfiler.py

Skriver `data/foia-svar/PROFIL.md` och samma sak till skärmen.
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
from pathlib import Path

# Hur många rader som visas rå från toppen. Kommunexporter lägger ofta en titel
# och ett par tomma rader före rubrikraden, så rubriken går inte att anta ligga
# först — den måste synas för att kunna pekas ut.
TOPPRADER = 8
EXEMPELRADER = 3
MAX_CELL = 40

# Kolumnnamn som brukar bära det loadern behöver. Bara en gissning som
# markeras i utdatat; människan bekräftar.
GISSNING = {
    "datum": ("datum", "fakturadatum", "bokforingsdatum", "bokföringsdatum", "period", "bokf"),
    "belopp": ("belopp", "summa", "amount", "kostnad", "netto", "exkl"),
    "leverantör": ("leverantor", "leverantör", "supplier", "motpart", "namn"),
    "orgnr": ("orgnr", "organisationsnummer", "org.nr", "leverantor_id", "leverantörsid"),
}


def kort(varde: object) -> str:
    text = "" if varde is None else str(varde).strip()
    text = re.sub(r"\s+", " ", text)
    return text[:MAX_CELL] + "…" if len(text) > MAX_CELL else text


def gissa(rubrik: str) -> str | None:
    låg = rubrik.casefold()
    for roll, nycklar in GISSNING.items():
        if any(n in låg for n in nycklar):
            return roll
    return None


def är_datum(varde: object) -> bool:
    if isinstance(varde, dt.date | dt.datetime):
        return True
    text = str(varde or "").strip()
    return bool(re.fullmatch(r"\d{4}[-/]?\d{2}[-/]?\d{2}([ T].*)?", text))


def profil_xlsx(sokvag: Path) -> list[str]:
    try:
        import openpyxl
    except ImportError:
        return ['  openpyxl saknas — `pip install -e ".[files]"`']

    ut: list[str] = []
    bok = openpyxl.load_workbook(sokvag, read_only=True, data_only=True)
    try:
        ut.append(f"  Blad: {', '.join(bok.sheetnames)}")
        for namn in bok.sheetnames:
            blad = bok[namn]
            ut.append(f"\n  ### Blad `{namn}`")
            rader = blad.iter_rows(values_only=True)

            topp = []
            for i, rad in enumerate(rader):
                if i >= TOPPRADER:
                    break
                topp.append(rad)
            if not topp:
                ut.append("  (tomt blad)")
                continue

            ut.append(f"  Första {len(topp)} raderna, för att hitta rubrikraden:")
            for i, rad in enumerate(topp):
                celler = " | ".join(kort(c) for c in rad[:12])
                ut.append(f"    rad {i}: {celler}")

            # Rubrikraden antas vara den första med flest ifyllda celler.
            rubrikindex = max(
                range(len(topp)), key=lambda i: sum(1 for c in topp[i] if c not in (None, ""))
            )
            rubriker = [kort(c) for c in topp[rubrikindex]]
            ut.append(f"\n  Trolig rubrikrad: rad {rubrikindex}")
            for j, r in enumerate(rubriker):
                if not r:
                    continue
                roll = gissa(r)
                ut.append(f"    [{j}] {r}" + (f"   <- ser ut som {roll}" if roll else ""))

            # Strömmande genomgång: radantal och datumspann per kolumn som ser
            # ut att bära datum. Detta är frågan som avgör om filen duger.
            antal = 0
            spann: dict[int, list[str]] = {}
            for rad in blad.iter_rows(min_row=rubrikindex + 2, values_only=True):
                if all(c in (None, "") for c in rad):
                    continue
                antal += 1
                if antal <= 20000:  # räcker för att fastställa spannets form
                    for j, c in enumerate(rad):
                        if är_datum(c):
                            text = (
                                c.isoformat()[:10]
                                if isinstance(c, dt.date | dt.datetime)
                                else str(c)[:10]
                            )
                            nuv = spann.setdefault(j, [text, text])
                            nuv[0] = min(nuv[0], text)
                            nuv[1] = max(nuv[1], text)
            ut.append(f"\n  Datarader: {antal}")
            if spann:
                for j, (lag, hog) in sorted(spann.items()):
                    rubrik = rubriker[j] if j < len(rubriker) else f"kolumn {j}"
                    ut.append(f"    datum i [{j}] {rubrik}: {lag} – {hog}")
            else:
                ut.append(
                    "    INGEN datumkolumn hittad. Utan fakturadatum per rad kan filen "
                    "inte användas för utnyttjandegrad — to_payments kastar raden."
                )
    finally:
        bok.close()
    return ut


def profil_xls(sokvag: Path) -> list[str]:
    try:
        import xlrd
    except ImportError:
        return ['  xlrd saknas — `pip install -e ".[files]"`']
    ut: list[str] = []
    bok = xlrd.open_workbook(sokvag)
    ut.append(f"  Blad: {', '.join(bok.sheet_names())}")
    for blad in bok.sheets():
        ut.append(f"\n  ### Blad `{blad.name}` — {blad.nrows} rader, {blad.ncols} kolumner")
        for i in range(min(TOPPRADER, blad.nrows)):
            celler = " | ".join(kort(blad.cell_value(i, j)) for j in range(min(12, blad.ncols)))
            ut.append(f"    rad {i}: {celler}")
    return ut


def profil_pdf(sokvag: Path) -> list[str]:
    try:
        from pypdf import PdfReader
    except ImportError:
        return ['  pypdf saknas — `pip install -e ".[files]"`']
    lasare = PdfReader(sokvag)
    ut = [f"  Sidor: {len(lasare.pages)}"]
    text = (lasare.pages[0].extract_text() or "").strip()
    if not text:
        ut.append("  Sidan 1 ger ingen text — troligen inskannad. OCR krävs.")
        return ut
    rader = [r for r in text.splitlines() if r.strip()][:20]
    ut.append("  Sidan 1, första raderna:")
    ut += [f"    {kort(r)}" for r in rader]
    ut.append(
        "  Bedöm om raderna är kolumnuppdelade. Är de det går tabellen att läsa; "
        "är de löpande text behövs ett riktigt tabellverktyg."
    )
    return ut


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--katalog", default="data/foia-svar", help="Var filerna ligger")
    args = parser.parse_args()

    katalog = Path(args.katalog)
    if not katalog.is_dir():
        print(f"{katalog} finns inte. Kör `python3 scripts/hamta_bilagor.py --live` först.")
        return 1

    filer = sorted(p for p in katalog.iterdir() if p.suffix.lower() in {".xlsx", ".xls", ".pdf"})
    if not filer:
        print(f"Inga xlsx/xls/pdf i {katalog}.")
        return 1

    rader = [
        "# Profil av utlämnade filer",
        "",
        f"_{len(filer)} filer i `{katalog}`. Struktur, inte innehåll — "
        "litet nog att klistra in i en session._",
        "",
    ]
    for sokvag in filer:
        mb = sokvag.stat().st_size / 1e6
        rader += [f"## `{sokvag.name}` ({mb:.1f} MB)", ""]
        try:
            match sokvag.suffix.lower():
                case ".xlsx":
                    rader += profil_xlsx(sokvag)
                case ".xls":
                    rader += profil_xls(sokvag)
                case ".pdf":
                    rader += profil_pdf(sokvag)
        except Exception as fel:  # noqa: BLE001 — en trasig fil ska inte stoppa de andra
            rader.append(f"  KUNDE INTE LÄSAS: {type(fel).__name__}: {fel}")
        rader.append("")

    text = "\n".join(rader)
    mal = katalog / "PROFIL.md"
    mal.write_text(text, encoding="utf-8")
    print(text)
    print(f"\n---\nSkrivet till {mal} ({len(text) / 1000:.1f} kB). Klistra in den i sessionen.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
